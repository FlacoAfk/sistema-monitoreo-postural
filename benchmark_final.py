"""
BENCHMARK FINAL — Evaluación definitiva de 10 modelos YOLO-Pose
Genera una carpeta única de pruebas + Excel con métricas completas.
Mapeo corregido: K0 (cabeza), K6 (C7), K7 (escápula) = nodos críticos.
"""
import cv2
import numpy as np
import os
import time
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ultralytics import YOLO

# === CONFIGURACIÓN ===
BASE = r"c:\Users\elkaw\Desktop\Modelos entrenados"
DATASET = r"C:\Users\elkaw\Desktop\Nueva carpeta"
OUT_DIR = os.path.join(BASE, "imagenes_prueba_postura")

KP_NAMES = {
    0: "K0-HeadBack",   1: "K1-NeckBack",    2: "K2-ShoulderTop",
    3: "K3-BackEdge", 4: "K4-HipsEdge",    5: "K5-NeckMid",
    6: "K6-Jaw",       7: "K7-Chin",   8: "K8-ShoulderBack",
}
CRITICAL_KPS = {0, 1, 3}

COLORS = {
    0: (255, 80, 80),   1: (80, 255, 80),   2: (80, 80, 255),
    3: (255, 255, 80),  4: (255, 80, 255),  5: (80, 255, 255),
    6: (0, 0, 255),     7: (0, 165, 255),   8: (255, 200, 80),
}

SKELETON = [
    (0, 1), (0, 2), (2, 6), (6, 7), (6, 5),
    (5, 3), (7, 4), (3, 8), (8, 4),
]

MODELS = [
    "yolov5n_pose_b16_lr05",
    "yolov5m_pose_b32_lr01",
    "yolov8n_pose_b16_lr05",
    "yolov8s_pose_b8_lr05",
    "yolov11n_pose_b16_lr01",
    "yolov11s_pose_b32_lr05",
    "yolov11m_pose_b64_lr01",
    "yolov26n_pose_b128_lr05",
    "yolov26s_pose_b64_lr03",
    "yolov26m_pose_b16_lr03",
]


def parse_gt(label_path, img_w, img_h):
    if not os.path.exists(label_path):
        return None
    with open(label_path, 'r') as f:
        lines = f.readlines()
    if not lines:
        return None
    parts = lines[0].strip().split()
    if len(parts) < 5 + 27:
        return None
    kps = []
    for k in range(9):
        base = 5 + k * 3
        kx = float(parts[base]) * img_w
        ky = float(parts[base + 1]) * img_h
        kv = float(parts[base + 2])
        kps.append([kx, ky, kv])
    return np.array(kps)


def compute_oks(gt_kps, pred_kps, bbox_area, kp_indices=None):
    if kp_indices is None:
        kp_indices = range(9)
    sigmas = np.array([0.05] * 9)
    oks_vals = []
    for i in kp_indices:
        if gt_kps[i][2] < 1:
            continue
        dx = gt_kps[i][0] - pred_kps[i][0]
        dy = gt_kps[i][1] - pred_kps[i][1]
        s = bbox_area
        oks_i = np.exp(-(dx**2 + dy**2) / (2 * s * sigmas[i]**2))
        oks_vals.append(oks_i)
    return np.mean(oks_vals) if oks_vals else 0.0


def draw_kps(img, kps, title, conf_thresh=0.3, is_gt=False):
    out = img.copy()
    h, w = out.shape[:2]
    
    k0_occ, k1_cerv, k3_dorsal = None, None, None
    
    for (a, b) in SKELETON:
        if a >= len(kps) or b >= len(kps):
            continue
        ca = kps[a][2] if len(kps[a]) > 2 else 1.0
        cb = kps[b][2] if len(kps[b]) > 2 else 1.0
        if is_gt or (ca > conf_thresh and cb > conf_thresh):
            xa, ya = int(kps[a][0]), int(kps[a][1])
            xb, yb = int(kps[b][0]), int(kps[b][1])
            if xa > 0 and ya > 0 and xb > 0 and yb > 0:
                color = (0, 140, 255) if (a in CRITICAL_KPS and b in CRITICAL_KPS) else (100, 100, 100)
                cv2.line(out, (xa, ya), (xb, yb), color, 2, cv2.LINE_AA)
    
    for i in range(min(len(kps), 9)):
        x, y = int(kps[i][0]), int(kps[i][1])
        conf = kps[i][2] if len(kps[i]) > 2 else 1.0
        if (not is_gt and conf < conf_thresh) or (x <= 0 and y <= 0):
            continue
        is_crit = i in CRITICAL_KPS
        r = 8 if is_crit else 5
        cv2.circle(out, (x, y), r, COLORS[i], -1, cv2.LINE_AA)
        if is_crit:
            cv2.circle(out, (x, y), r + 3, (255, 255, 255), 2, cv2.LINE_AA)
        
        label = f"K{i}"
        font = cv2.FONT_HERSHEY_SIMPLEX
        (tw, th), _ = cv2.getTextSize(label, font, 0.45, 1)
        if i == 0:
            k0_occ = (x, y)
        elif i == 1:
            k1_cerv = (x, y)
        elif i == 3:
            k3_dorsal = (x, y)
        lx = x + 14 if i % 2 == 0 else x - tw - 14
        ly = y + 6
        if lx < 0: lx = x + 14
        if lx + tw > w: lx = x - tw - 14
        cv2.rectangle(out, (lx-2, ly-th-3), (lx+tw+2, ly+4), (0,0,0), -1)
        cv2.putText(out, label, (lx, ly), font, 0.45, COLORS[i], 1, cv2.LINE_AA)
        
    if k0_occ and k1_cerv and k3_dorsal:
        cv2.line(out, k1_cerv, k0_occ, (0, 255, 255), 2)  # Vector craneocervical
        cv2.line(out, k1_cerv, k3_dorsal, (0, 255, 255), 2)  # Vector dorsolumbar
    
    cv2.putText(out, title, (5, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 1, cv2.LINE_AA)
    return out


# === PREPARAR CARPETA DE SALIDA ===
print("🗑️  Limpiando carpeta de pruebas...")
if os.path.exists(OUT_DIR):
    shutil.rmtree(OUT_DIR)
os.makedirs(OUT_DIR)

# === SELECCIONAR 25 IMÁGENES VARIADAS ===
val_imgs_dir = os.path.join(DATASET, "valid", "images")
val_lbls_dir = os.path.join(DATASET, "valid", "labels")
all_val = sorted(os.listdir(val_imgs_dir))

preferred = []
for f in all_val:
    fl = f.lower()
    if any(kw in fl for kw in [
        "good_posture_1_", "good_posture_12", "good_posture_21", "good_posture_22",
        "good_posture_23", "good_posture_3_", "good_posture_5_",
        "bad_posture_6_", "bad_posture_10", "bad_posture_14",
        "bad_posture_2_", "bad_posture_3_", "bad_posture_5_",
    ]):
        preferred.append(f)

np.random.seed(42)
uniform = list(np.random.choice(all_val, 15, replace=False))
test_images = list(dict.fromkeys(preferred + uniform))[:25]

# Copiar las imágenes originales
originals_dir = os.path.join(OUT_DIR, "imagenes_originales")
os.makedirs(originals_dir)
for img_name in test_images:
    src = os.path.join(val_imgs_dir, img_name)
    dst = os.path.join(originals_dir, img_name)
    shutil.copy2(src, dst)
print(f"📸 {len(test_images)} imágenes copiadas a imagenes_originales/")

# === BENCHMARK POR MODELO ===
all_results = []

for model_name in MODELS:
    weights = os.path.join(BASE, model_name, "weights", "best.pt")
    if not os.path.exists(weights):
        print(f"⚠️  {model_name}: no encontrado")
        continue
    
    print(f"\n🔄 Evaluando {model_name}...")
    model = YOLO(weights)
    
    model_dir = os.path.join(OUT_DIR, model_name)
    os.makedirs(model_dir)
    
    times = []
    detections = 0
    oks_all_list = []
    oks_crit_list = []
    dist_per_kp = {i: [] for i in range(9)}
    visual_pass = 0
    visual_total = 0
    
    for img_name in test_images:
        img_path = os.path.join(val_imgs_dir, img_name)
        img = cv2.imread(img_path)
        if img is None:
            continue
        h, w = img.shape[:2]
        
        lbl_name = os.path.splitext(img_name)[0] + ".txt"
        lbl_path = os.path.join(val_lbls_dir, lbl_name)
        gt_kps = parse_gt(lbl_path, w, h)
        
        # Inferencia
        t0 = time.perf_counter()
        res = model.predict(img, verbose=False, conf=0.25)
        t1 = time.perf_counter()
        times.append((t1 - t0) * 1000)
        
        result = res[0]
        pred_kps = None
        if result.keypoints is not None and len(result.keypoints) > 0:
            best_idx = result.boxes.conf.argmax().item()
            pred_kps = result.keypoints.data[best_idx].cpu().numpy()
            detections += 1
        
        # Métricas
        if gt_kps is not None and pred_kps is not None:
            gt_xs = [gt_kps[i][0] for i in range(9) if gt_kps[i][2] > 0]
            gt_ys = [gt_kps[i][1] for i in range(9) if gt_kps[i][2] > 0]
            bbox_area = max((max(gt_xs)-min(gt_xs)) * (max(gt_ys)-min(gt_ys)), 1) if gt_xs else w*h
            
            oks_all_list.append(compute_oks(gt_kps, pred_kps, bbox_area))
            oks_crit_list.append(compute_oks(gt_kps, pred_kps, bbox_area, CRITICAL_KPS))
            
            for i in range(9):
                if gt_kps[i][2] > 0 and pred_kps[i][2] > 0.3:
                    dist = np.sqrt((gt_kps[i][0]-pred_kps[i][0])**2 + (gt_kps[i][1]-pred_kps[i][1])**2)
                    dist_per_kp[i].append(dist)
            
            # Prueba visual: los 3 críticos con <20px de error
            visual_total += 1
            crit_dists = []
            for ci in CRITICAL_KPS:
                if gt_kps[ci][2] > 0 and pred_kps[ci][2] > 0.3:
                    d = np.sqrt((gt_kps[ci][0]-pred_kps[ci][0])**2 + (gt_kps[ci][1]-pred_kps[ci][1])**2)
                    crit_dists.append(d)
            if crit_dists and all(d < 20 for d in crit_dists):
                visual_pass += 1
        
        # Generar imagen comparativa GT vs Predicción
        gt_vis = draw_kps(img, gt_kps, "GT (Ground Truth)", is_gt=True) if gt_kps is not None else img.copy()
        pred_vis = draw_kps(img, pred_kps, f"{model_name}", conf_thresh=0.3) if pred_kps is not None else img.copy()
        combined = np.hstack([gt_vis, pred_vis])
        short = os.path.splitext(img_name)[0][:40]
        cv2.imwrite(os.path.join(model_dir, f"{short}.jpg"), combined, [cv2.IMWRITE_JPEG_QUALITY, 90])
    
    # Estadísticas
    avg_time = np.mean(times) if times else 0
    med_time = np.median(times) if times else 0
    min_time = np.min(times) if times else 0
    det_rate = detections / len(test_images) * 100
    avg_oks = np.mean(oks_all_list) if oks_all_list else 0
    avg_oks_crit = np.mean(oks_crit_list) if oks_crit_list else 0
    visual_rate = (visual_pass / visual_total * 100) if visual_total > 0 else 0
    
    avg_dist = {}
    for i in range(9):
        avg_dist[i] = np.mean(dist_per_kp[i]) if dist_per_kp[i] else -1
    
    arch = model_name.split("_")[0]
    size_code = model_name.split("_")[0][-1]
    size = {"n": "nano", "s": "small", "m": "medium"}.get(size_code, "?")
    
    row = {
        "modelo": model_name,
        "arquitectura": arch,
        "tamaño": size,
        "avg_ms": round(avg_time, 1),
        "med_ms": round(med_time, 1),
        "min_ms": round(min_time, 1),
        "fps_estimado": round(1000 / avg_time, 1) if avg_time > 0 else 0,
        "det_rate": round(det_rate, 1),
        "oks_global": round(avg_oks, 4),
        "oks_criticos": round(avg_oks_crit, 4),
        "dist_K0_px": round(avg_dist[0], 1) if avg_dist[0] >= 0 else "N/A",
        "dist_K1_px": round(avg_dist[1], 1) if avg_dist[1] >= 0 else "N/A",
        "dist_K3_px": round(avg_dist[3], 1) if avg_dist[3] >= 0 else "N/A",
        "dist_avg_px": round(np.mean([v for v in avg_dist.values() if v >= 0]), 1),
        "visual_pass": f"{visual_pass}/{visual_total}",
        "visual_rate": round(visual_rate, 1),
    }
    all_results.append(row)
    
    print(f"   ⏱️  {avg_time:.0f}ms ({1000/avg_time:.0f}fps) | Det: {det_rate:.0f}% | OKS★: {avg_oks_crit:.3f} | Visual: {visual_pass}/{visual_total}")

# === RANKING ===
max_time = max(r["avg_ms"] for r in all_results)
for r in all_results:
    speed_score = 1 - (r["avg_ms"] / max_time)
    det_score = r["det_rate"] / 100
    r["score"] = round(r["oks_criticos"] * 0.40 + r["oks_global"] * 0.20 + speed_score * 0.20 + det_score * 0.20, 4)

all_results.sort(key=lambda x: x["score"], reverse=True)

for i, r in enumerate(all_results):
    r["ranking"] = i + 1

# === GENERAR EXCEL ===
print("\n📊 Generando Excel...")
wb = Workbook()

# --- Hoja 1: Ranking ---
ws1 = wb.active
ws1.title = "Ranking"

# Estilos
gold = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
silver = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
bronze = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
fourth = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
bold_font = Font(bold=True, size=11)
normal_font = Font(size=10)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

headers = [
    "#", "Modelo", "Arquitectura", "Tamaño", 
    "Vel. Prom (ms)", "Vel. Med (ms)", "FPS Est.",
    "Det %", "OKS Global", "OKS Críticos ★",
    "Dist K0 (px)", "Dist K1 (px)", "Dist K3 (px)", "Dist Prom (px)",
    "Prueba Visual", "% Visual OK", "SCORE"
]

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

for row_idx, r in enumerate(all_results, 2):
    vals = [
        r["ranking"], r["modelo"], r["arquitectura"], r["tamaño"],
        r["avg_ms"], r["med_ms"], r["fps_estimado"],
        r["det_rate"], r["oks_global"], r["oks_criticos"],
        r["dist_K0_px"], r["dist_K1_px"], r["dist_K3_px"], r["dist_avg_px"],
        r["visual_pass"], r["visual_rate"], r["score"]
    ]
    
    fill = None
    if r["ranking"] == 1: fill = gold
    elif r["ranking"] == 2: fill = silver
    elif r["ranking"] == 3: fill = bronze
    elif r["ranking"] == 4: fill = fourth
    
    for col, val in enumerate(vals, 1):
        cell = ws1.cell(row=row_idx, column=col, value=val)
        cell.font = bold_font if r["ranking"] <= 4 else normal_font
        cell.alignment = center
        cell.border = thin_border
        if fill:
            cell.fill = fill

# Ajustar anchos
col_widths = [4, 32, 14, 10, 14, 14, 10, 8, 12, 14, 12, 12, 12, 14, 12, 12, 10]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# --- Hoja 2: Mapeo de Keypoints ---
ws2 = wb.create_sheet("Mapeo Keypoints")
kp_headers = ["K (YOLO)", "ID Roboflow", "Descripción anatómica", "Nodo crítico"]
for col, h in enumerate(kp_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

kp_data = [
    ("K0", 0, "Cabeza superior frontal — Coronilla", "SÍ ★"),
    ("K1", 1, "Mentón / Mandíbula", ""),
    ("K2", 2, "Occipital — Parte posterior cabeza", ""),
    ("K3", 6, "Borde anterior del torso — Pecho", ""),
    ("K4", 7, "Cadera / Lumbosacra — Punto más bajo", ""),
    ("K5", 10, "Acromion — Hombro superior", ""),
    ("K6", 13, "Cervical posterior (C7) — Pivote del ángulo θ", "SÍ ★ (pivote)"),
    ("K7", 14, "Borde posterior dorsal / Escápula", "SÍ ★"),
    ("K8", 18, "Borde anterior pectoral", ""),
]
crit_fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
for row_idx, (k, rf_id, desc, crit) in enumerate(kp_data, 2):
    vals = [k, rf_id, desc, crit]
    for col, val in enumerate(vals, 1):
        cell = ws2.cell(row=row_idx, column=col, value=val)
        cell.font = normal_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        if crit:
            cell.fill = crit_fill

ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 45
ws2.column_dimensions['D'].width = 20

# --- Hoja 3: Metodología ---
ws3 = wb.create_sheet("Metodologia")
info = [
    ("Campo", "Detalle"),
    ("Dataset", "Desk Posture v2 (Roboflow) — 21,262 imágenes"),
    ("Imágenes evaluadas", f"{len(test_images)} del set de validación"),
    ("Nodos críticos", "K0 (Occipital), K1 (C7 cervical), K3 (Borde dorsal)"),
    ("Ángulo θ", "θ = ∠(K1→K0, K1→K3) — flexión cervicodorsal"),
    ("Métrica OKS", "Object Keypoint Similarity (σ=0.05 uniforme)"),
    ("Prueba visual", "PASA si los 3 keypoints críticos están a <20px del GT"),
    ("Score compuesto", "OKS_crit×40% + OKS_global×20% + Velocidad×20% + Detección×20%"),
    ("Fecha evaluación", "2026-05-07"),
    ("Nota", "Los IDs de Roboflow (0,1,2,6,7,10,13,14,18) se remapean a 0-8 en YOLO"),
]
for row_idx, (campo, detalle) in enumerate(info, 1):
    cell_a = ws3.cell(row=row_idx, column=1, value=campo)
    cell_b = ws3.cell(row=row_idx, column=2, value=detalle)
    if row_idx == 1:
        cell_a.font = header_font
        cell_a.fill = header_fill
        cell_b.font = header_font
        cell_b.fill = header_fill
    else:
        cell_a.font = Font(bold=True, size=10)
        cell_b.font = normal_font
    cell_a.border = thin_border
    cell_b.border = thin_border

ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 65

# Guardar
excel_path = os.path.join(OUT_DIR, "benchmark_modelos_postura.xlsx")
wb.save(excel_path)

# === IMPRIMIR RANKING ===
print("\n" + "=" * 90)
print("🏆 RANKING FINAL")
print("=" * 90)
print(f"{'#':<3} {'Modelo':<30} {'ms':>5} {'FPS':>5} {'Det%':>5} {'OKS★':>6} {'K0px':>5} {'K1px':>5} {'K3px':>5} {'Vis%':>5} {'SCORE':>7}")
print("-" * 90)
for r in all_results:
    m = {1:"🥇", 2:"🥈", 3:"🥉", 4:" 4"}.get(r["ranking"], "  ")
    k0 = r["dist_K0_px"] if isinstance(r["dist_K0_px"], (int,float)) else 0
    k1 = r["dist_K1_px"] if isinstance(r["dist_K1_px"], (int,float)) else 0
    k3 = r["dist_K3_px"] if isinstance(r["dist_K3_px"], (int,float)) else 0
    print(f"{m} {r['modelo']:<30} {r['avg_ms']:>5.0f} {r['fps_estimado']:>5.0f} {r['det_rate']:>5.0f} {r['oks_criticos']:>.4f} {k0:>5.1f} {k1:>5.1f} {k3:>5.1f} {r['visual_rate']:>5.0f} {r['score']:>7.4f}")

print(f"\n✅ Excel: {excel_path}")
print(f"✅ Carpeta: {OUT_DIR}")
print(f"   📁 imagenes_originales/ — {len(test_images)} imágenes de prueba")
for m in MODELS:
    print(f"   📁 {m}/ — comparaciones GT vs Predicción")
